import openpyxl as excel

book =excel.Workbook()
sheet =book.active
sheet["A1"] ="勤勉な人の計画は成功する"
sheet.cell(row=2, column=1, value="サルの尻笑い")
cell =sheet.cell(row=3, column=1)
cell.value ="捜すのに時がありあきらめる時がある"

book.save("cell_w.xlsx")
